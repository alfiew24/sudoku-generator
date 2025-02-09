from random import randint
from math import floor
from openpyxl import load_workbook

# n by n size
n = 9

board = [[] for y in range(n)]

class pallette:
    def __init__(self, value, options): # Options are stored in nested lists going U-R-D-L
        self.value = value
        self.up_options = options[0]
        self.right_options = options[1]
        self.down_options = options[2]
        self.left_options = options[3]

class tile:
    def __init__(self):
        self.options = {'1', '2', '3', '4', '5', '6', '7', '8', '9'}
        r = 0
        while len(board[r]) == n:
            r += 1

        self.x, self.y = r, len(board[r]) # x is row and y is column - we can find it by board[x][y]
        board[r].append(self)
        

    def collapse(self, choice):

        self.options = {choice}

        for i in range(9):
            if i != self.y and choice in board[self.x][i].options:
                board[self.x][i].options.remove(choice)
                # If collapsed to singular value, then collapse this cell
                if len(board[self.x][i].options) == 1:
                    board[self.x][i].collapse(list(board[self.x][i].options)[0])
            if i != self.x and choice in board[i][self.y].options:
                board[i][self.y].options.remove(choice)
                if len(board[i][self.y].options) == 1:
                    board[i][self.y].collapse(list(board[i][self.y].options)[0])

        chunk_x = floor(self.x / 3.0)
        chunk_y = floor(self.y / 3.0)
        
        for i in range(3):
            for j in range(3):
                if (chunk_x*3 + i != self.x or chunk_y*3 + j != self.y) and choice in board[chunk_x*3 + i][chunk_y*3 + j].options:
                        board[chunk_x*3 + i][chunk_y*3 + j].options.remove(choice)
                        if len(board[chunk_x*3 + i][chunk_y*3 + j].options) == 1:
                            board[chunk_x*3 + i][chunk_y*3 + j].collapse(list(board[chunk_x*3 + i][chunk_y*3 + j].options)[0])




def show_board():
    print('- - - - - - -')
    for row in board:
        for t in row:
            print(t.options, end='\t')
        print('\n')
    print('- - - - - - -')

def lowest_entropy():
    entropies = []
    for row in board:
        r = []
        for t in row:
            if len(t.options) > 1:
                r.append(len(t.options))
        if len(r) > 0:
            entropies.append(min(r))

    choices = []
    for row in board:
        for t in row:
            if len(t.options) == min(entropies):
                choices.append(t)

    return choices[randint(0, len(choices) - 1)]
    



for i in range(int(n*n)):
    tile()

board[randint(0, 8)][randint(0, 8)].collapse('1')

while True:
    try:
        choice = lowest_entropy()
        choice.collapse(list(choice.options)[randint(0, len(choice.options)-1)])
    except:
        break

show_board()
file = load_workbook('Sudoku Puzzle.xlsx')
for row in range(len(board)):
    for cell in range(len(board[row])):
        file['Solution'].cell(row=row + 1, column=cell + 1).value = int(list(board[row][cell].options)[0])
        
        file['Puzzle'].cell(row=row + 1, column=cell + 1).value = ''
        if randint(1, 10) > 5:
            file['Puzzle'].cell(row=row + 1, column=cell + 1).value = int(list(board[row][cell].options)[0])

file.save('Sudoku Puzzle.xlsx')
